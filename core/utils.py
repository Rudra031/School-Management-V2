import csv
import io
from django.http import HttpResponse
from django.utils import timezone
from core.models import AuditLog

def log_audit(request, action, module, model_name='', object_id='', object_repr='', changes=None):
    """
    Helper function to record an immutable AuditLog entry from anywhere in the codebase.
    """
    try:
        user = request.user if request and request.user.is_authenticated else None
        ip_address = getattr(request, 'client_ip', None)
        user_agent = getattr(request, 'user_agent', '')
        
        AuditLog.objects.create(
            user=user,
            action=action,
            module=module,
            model_name=model_name,
            object_id=str(object_id),
            object_repr=str(object_repr)[:255],
            changes=changes or {},
            ip_address=ip_address,
            user_agent=user_agent[:500] if user_agent else '',
        )
    except Exception as e:
        # Logging failure should never crash the primary business transaction
        pass


def export_to_csv(filename, headers, rows):
    """
    Generates a downloadable CSV HTTP response.
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


def export_to_excel(filename, sheet_name, headers, rows):
    """
    Generates a professional XLSX Excel file response using openpyxl,
    or falls back to CSV if openpyxl is unavailable.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        # Header styling
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 25

        # Data rows
        for row_idx, row_data in enumerate(rows, start=2):
            ws.append(row_data)
            for col_num in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.font = Font(name='Arial', size=10)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        return response
    except ImportError:
        return export_to_csv(filename, headers, rows)
